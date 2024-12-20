import streamlit as st
from streamlit_extras.stylable_container import stylable_container
import pandas as pd
import numpy as np
from statistics import mean
from datetime import datetime
from meteostat import Hourly, Stations
from sankeyflow import Sankey
import matplotlib.pyplot as plt
import io
import pydeck as pdk
from geopy.geocoders import Nominatim
from io import BytesIO
from pyxlsb import open_workbook as open_xlsb
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows 


# Page setting
st.set_page_config(page_title="Gigafactory Builder",
                   layout="wide",
                   page_icon="Giga_Builder_logo_small.png")

with open('style.css') as f:
    st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)

st.logo("Gigafactory Builder Logo.png", size="large")


#-----SIDEBAAARRR----------------------------------------------------------

#-----INFO BUTTON-----------------------------------
with st.sidebar.expander("**:material/info: Introduction**"):
    st.markdown("**Welcome to the Gigafactory Builder.** The Gigafactory Builder is an interactive planning tool for battery cell factories which is designed to support planners, decision-makers, and other stakeholders in early planning phases. It provides a scientifically sound, neutral basis for decision-making.")
    st.markdown("The Fraunhofer Research Institution for Battery Cell Production FFB offers this tool, which is based on the latest findings and calculations as well as a deep understanding of the processes in battery cell production.")
    st.markdown("If you have any questions, please feel free to contact us and we will try to find an answer or a solution together.")

st.sidebar.header('selectable planning parameters')

#-----GET GEOPY LOCATION COORDINATES-----------------------------------------------------------
# ... location input ...
location_geopy= st.sidebar.text_input("**:material/location_on: location**","Münster", help="Sets the location of your gigafactory. The climate of the location can drastically change the energy demand of the factory.")

# ... find location coordinates via geocode ...
geolocator = Nominatim(user_agent="user_agent")
location = geolocator.geocode(f"{location_geopy}",exactly_one=True, language="en", namedetails=True, addressdetails=True)

#-----COUNTRY CODE---------------------------------
country_code = location.raw['address']['country_code']
st.sidebar.write(location.address)


lat = location.latitude
lon = location.longitude



#-----Display the map----------------------------------------------------------------------
center_lng = lon
center_lat = lat

# Define view_state
view_state = pdk.ViewState(
    latitude=center_lat, longitude=center_lng, zoom=10
)
# Define the icon mapping

IMG_URL = "https://upload.wikimedia.org/wikipedia/commons/3/3b/Battery.svg"

#-----make coordinates df------------------------------------------
d = [lon,lat]
coordinates_df = pd.DataFrame(np.array([d]), columns=["longitude", "latitude"])

# Create the layer with the custom symbol
icon = pdk.Layer(
    type="BitmapLayer",
    data=coordinates_df,
    get_position=["longitude", "latitude"],
    image=IMG_URL
)

# Create the map
map = pdk.Deck(
    initial_view_state=view_state,
    layers=[icon],
    map_style=f"mapbox://styles/mapbox/{'streets-v11'}",
)

st.sidebar.pydeck_chart(map, height=250)


#----------------------------------------------------------------------------------------------------------

with st.sidebar.container(border=True):
    production_capacity= st.sidebar.slider('**production capacity [GWh/a]**', 5, 150, 40, help="Determines how much battery capacity will be produced in your factory per year. This value assumes 315 production days. If you change the production days, you will see the actual value change in the blue box. :D")
cell_format = st.sidebar.selectbox('**cell format**', ('Pouch', 'Cylindrical', 'Prismatic'), help="Different cell formats have different manufacturing steps and requirements. Choose between the three most common formats.")
automation_degree = st.sidebar.selectbox('**degree of automation**',('low','normal','high'), help="Depending on the degree of automation, more or less people are working in the dry rooms, which also has an impact on energy demand.")
dew_point = st.sidebar.selectbox('**dew point in dry rooms**',('-60 °C','-50 °C','-40 °C'), help="The dew point in the dry rooms are different for different cell chemistries.")
production_days = st.sidebar.slider('**production days per year**', 1, 365, 315, help="How many days is your factory running on full manufacturing capacity? If you set this to less than 315 days, the production capacity has to be seen as theoretical, as there of course will be less battery cells produced if you have fewer production days.")
energy_concept = st.sidebar.selectbox('**energy concept**', ('Natural Gas Boiler', 'Cogeneration Unit', 'Heat Pump', 'Hybrid Heat Pump'), help="The energy concept defines where the energy is coming from. Depending on your choice, the heat output is generated with electricity through a heat pump, or with natural gas.") 
st.sidebar.subheader('Developer Options')
year = st.sidebar.slider('**weather reference year**', 2003, 2023, 2023, help="Experimental feature that changes the reference year the Gigafactory Builder uses to calculate the energy demand of the process steps depending on outside temperature.")
electricity_price=st.sidebar.slider('Electricity Price in €/kWh',0.05,0.50,0.15, help="The price of electricity determines the cost of energy.")


st.sidebar.markdown('''
---
Created by Fraunhofer FFB
''')

st.sidebar.link_button("Fraunhofer FFB","https://www.ffb.fraunhofer.de/",use_container_width=True)

#----input for displayed production capacity-----------------------------
production_day_factor_315 = production_days/315

#Ausgabe vom Standort - nach Stadt - Bundesland - Staat -Reihenfolge
if location and 'city' in location.raw['address']:
    city = location.raw['address']['city']
else:
    if location and "state" in location.raw["address"]:
        city = location.raw['address']['state']
    else: 
        if location and "country" in location.raw["address"]:
            city = location.raw['address']['country']
        else: 
            city = "undefined"
#-----HEADER------------------------------------------------------------------
header_container = st.container(border=True)
with header_container:
    header1, header2 = st.columns([2,5])
    with header1:
        with stylable_container(
                key="top_battery",
                css_styles="""
                    button {
                        background-color: #83d1a1;
                        padding: 2% 2% 2% 2%;
                        border-width: 5px;
                        border-radius: 20px;
                    }
                    """
            ):
            st.image("GigaFactory_Builder_Logo_Entwurf.png")
    with header2:
        with stylable_container(
                key="top_battery",
                css_styles="""
                    button {
                        background-color: #83d1a1;
                        padding: 2% 2% 2% 2%;
                        border-width: 5px;
                        border-radius: 20px;
                    }
                    """
            ):
            st.subheader(":material/battery_charging_full: Your Factory", help="This area contains the basic information about your battery cell gigafactory based on the input in the sidebar.")
            
            battery5, battery6 = st.columns(2)
            with battery5:
                st.metric(label=":material/conveyor_belt: Location", value=f"{city}", help="The location of you factory based on your input.")
            with battery6:
                st.metric(label=":material/battery_unknown: Cell Format", value=cell_format, help="The cell format you chose.")
            battery1, battery2, battery3 = st.columns(3)
            with battery1:
                st.metric(label=":material/conveyor_belt: Actual Production Capacity [GWh/a]", value=round((production_capacity*production_day_factor_315),2), help="The actual annual production capacity in GWh/a. As you might have noticed, the production capacity you set in the sidebar refers to 315 production days. If you set your number lower or higher than that, this metric displays your actual production capacity that is used to calculate the values below.")
            with battery2:
                st.metric(label=":material/calendar_month: Production Days", value=production_days, help="The number of annual production days you selected.")
            with battery3:
                st.metric(label=":material/dew_point: Dew Point Temperature [°C]", value=dew_point, help="The dew point temperature in °C you set.")

#-----popover---------------------------------------------------


#---------------------------WEATHER DATA-----------------------------------------------
#----start_date----------------------
def get_start_dates(date):
    dates_dict = {}
    for year in range(date, 1989, -1):  # Rückwärts iterieren von 'date' bis 1990
        dates_dict[str(year)] = (year, 1, 1)
    return dates_dict.get(str(date), (2023,1,1))

#-----end date------------------------
def get_end_dates(date):
    dates_dict = {}
    for year in range(date, 1989, -1):  # Rückwärts iterieren von 'date' bis 1990
        dates_dict[str(year)] = (year, 12, 31)
    return dates_dict.get(str(date), (2023,12,31))

#-----get dates from chosen year---------------------------------
y1, m1, d1 = get_start_dates(year)
y2, m2, d2 = get_end_dates(year)




#-----GET WEATHER DATA---------------------------------------
stations = Stations()
stations = stations.nearby(lat, lon)
station = stations.fetch(1)
station['id']=station.index
stat=[]
for sta in range(1):
    stat.append([station['name'][sta],station['id'][sta]])

#-----weather param humidity---------------------------------
weather_param = 'rhum'

ypoints = []
for station_name, station_id in stat:
  # Download hourly data for a specific date range (adjust dates as needed)
  start_date = datetime(y1, m1, d1)  # Change year, month, day as needed
  end_date = datetime(y2, m2, d2)  # Change year, month, day as needed
  data = Hourly(f"{station_id}", start_date, end_date)
  data = data.fetch()

  station_rhum_data = data[weather_param].tolist()
  ypoints.append(station_rhum_data)

#-----weather param temperature---------------------------------
weather_param = 'temp'

ypoints = []
for station_name, station_id in stat:
  # Download hourly data for a specific date range (adjust dates as needed)
  start_date = datetime(y1, m1, d1)  # Change year, month, day as needed
  end_date = datetime(y2, m2, d2)  # Change year, month, day as needed
  data = Hourly(f"{station_id}", start_date, end_date)
  data = data.fetch()

  station_temp_data = data[weather_param].tolist()
  modified_data = []  # Create a new list to store modified temperature values
  for temp in station_temp_data:
    kelvin_temp = temp + 273.15
    modified_data.append(kelvin_temp)

#-----weather param pressure------------------------------------
weather_param = 'pres'

ypoints = []
for station_name, station_id in stat:
  # Download hourly data for a specific date range (adjust dates as needed)
  start_date = datetime(y1, m1, d1)  # Change year, month, day as needed
  end_date = datetime(y2, m2, d2)  # Change year, month, day as needed
  data = Hourly(f"{station_id}", start_date, end_date)
  data = data.fetch()

  station_pres_data = data[weather_param].tolist()
  ypoints.append(station_pres_data)






#---------------------------DEFINING LISTS---------------------------------------------
t = [float(value) for value in station_temp_data]
rhum = [float(value) for value in station_rhum_data]
pres = [float(value) for value in station_pres_data]
t_kelvin = [float(value) for value in modified_data]
f_abs = []
cool_data = []
heat_data = []
heat_data_partial = []
cool_data_partial = []
cop_data = []
eert_data = []
strom_wp_k_end = []
strom_wp_w_end = []
brennstoff_w_end = []
strom_electr_end = []


# make list with hours--------------------------------------
station_time_data = list(range(len(t)))


#---------------------------FORMULAS----------------------------------------------------------------------------
production_day_factor = production_days/365
standby_day_factor = (365-production_days)/365

#-----Anschlussleistung (Demo für Finn)-------------------------
def Anschlussleistung(x):
    return x*6.25

#-----RLT/HVAC Energieverbrauch-------------------------------------------------
#-----RLT Kältelast---------------------------------------------
def RLT_Kaeltelast(x):
    return 1.920*x
    
#-----RLT Wärmelast---------------------------------------------
def RLT_Waermelast(x):
    return 0.6867*x

#-----RLT Stromlast---------------------------------------------
def RLT_Stromlast(x):
    return 1.529*x


#-----MITARBEITENDE-----------------------------------------------------------
#MA pro GWh!!!!!!!!!!!!!!!!!!!!!
def MA_pro_GWh(x):
    return x

# a3.metric("Mitarbeitende im Trockenraum", f"{round(MA_nach_Automatisierungsgrad(MA_in_RuT(production_capacity, cell_format)),0)} MA")
#MA in RuT nach Produktionskapazität-------------------------
def MA_in_RuT(x, cell_format):
    factors = {
        "Pouch": 0.83,
        "Cylindrical": 1.0,
        "Prismatic": 1.225
    }
    return x * 27 * factors.get(cell_format, 1.0)

#MA Umrechnung nach Automatisierungsgrad--------------------
def MA_nach_Automatisierungsgrad(x2):
    if automation_degree == 'low':
        return x2*1.2
    if automation_degree == 'normal':
        return x2
    if automation_degree == 'high':
        return x2*0.8
    else:
        return x2

#-----PROZESSENERGIE--------------------------------------------------------
#-----Elektrische Last--------------------------------------------
def Prozess_Stromnutzlast(x):
    day_factor=365/315
    if cell_format == 'Pouch':
        return 25.86493*x*day_factor
    if cell_format == 'Cylindrical':
        return 26.59484*x*day_factor
    if cell_format == 'Prismatic':
        return 29.58601*x*day_factor
    
#-----Kälte-Nutzlast--------------------------------------------
def Prozess_Kaeltenutzlast(x):
    day_factor=365/315
    if cell_format == 'Pouch':
        return 8.14149*x*day_factor
    if cell_format == 'Cylindrical':
        return 9.60784*x*day_factor
    if cell_format == 'Prismatic':
        return 13.00309*x*day_factor


#-----REIN-UND TROCKENRAUM--------------------------------------------------
#-----absolute Feuchte----------------------------------------
def hum_abs(temp_val, rhum_val, pres_val):
    K1 = 6.112
    K2 = 17.62
    K3 = 243.12
    
    p_s = K1*np.exp(K2*temp_val/(K3+temp_val))*(rhum_val/100)
    x = 1000*(p_s/(pres_val-p_s)*0.622)
    if x>7.6:
        return(7.6)
    else:
        return(float(x))

#-----FULL HEAT DRY ROOM--------------------------------------
if dew_point == "-60 °C":
    def heat_full_dry_room(x, y):
        
        p00 =       103.3
        p10 =       -1.09
        p01 =      0.7667
        p20 =     0.02341
        p11 =    -0.07383
        p02 =     -0.1225
        p30 =   0.0006669
        p21 =   -0.001092
        p12 =     0.02655
        p03 =    -0.02426
        p40 =  -1.955e-05
        p31 =   8.267e-05
        p22 =   -0.000672
        p13 =   0.0004604
        p04 =   0.0007999
        return(p00+p10*x+p01*y+p20*x**2+p11*x*y+p02*y**2+p30*x**3+p21*x**2*y+p12*x*y**2+p03*y**3+p40*x**4+p31*x**3*y+p22*x**2*y**2+p13*x*y**3+p04*y**4)

#3-Rotor-System
if dew_point == "-50 °C":
    def heat_full_dry_room(x,y): 
        p00 =       21.58
        p10 =     -0.3012
        p01 =      0.2154
        p20 =    0.006439
        p11 =    -0.01973
        p02 =    -0.03474
        p30 =   0.0001893
        p21 =  -0.0003651
        p12 =    0.007437
        p03 =   -0.006864
        p40 =  -5.512e-06
        p31 =   2.409e-05
        p22 =   -0.000188
        p13 =   0.0001284
        p04 =    0.000232
        return(p00 + p10*x + p01*y + p20*x**2 + p11*x*y + p02*y**2 + p30*x**3 +p21*x**2*y + p12*x*y**2 + p03*y**3 + p40*x**4 + p31*x**3*y + p22*x**2*y**2 + p13*x*y**3 + p04*y**4)
#3-Rotor-System


if dew_point == "-40 °C":
    def heat_full_dry_room(x, y): 
        return ( 6.961 - 0.111 * x + 0.0695 * y + 0.002363 * x**2 - 0.007255 * x * y - 0.01146 * y**2 + 7.129e-05 * x**3 - 0.0001515 * x**2 * y + 0.002808 * x * y**2 - 0.003271 * y**3 - 2.057e-06 * x**4 + 9.04e-06 * x**3 * y - 6.803e-05 * x**2 * y**2 + 3.704e-05 * x * y**3 + 0.0001348 * y**4 )
#2-Rotor-System

#-----FULL COOL DRY ROOM--------------------------------------
if dew_point == "-60 °C":
    def cool_full_dry_room(x,y):
        
        p00 =        72.5
        p10 =     0.06158
        p01 =       2.406
        p20 =     0.02429
        p11 =    -0.09524
        p02 =      0.4824
        p30 =   0.0007075
        p21 =   -0.001683
        p12 =     0.03358
        p03 =     -0.1349
        p40 =  -2.039e-05
        p31 =   8.665e-05
        p22 =  -0.0006569
        p13 =   1.204e-05
        p04 =    0.007552

        return(p00 + p10*x + p01*y + p20*x**2 + p11*x*y + p02*y**2 + p30*x**3 + p21*x**2*y + p12*x*y**2 + p03*y**3 + p40*x**4 + p31*x**3*y + p22*x**2*y**2 + p13*x*y**3 + p04*y**4)
#3Rotor-System

if dew_point == "-50 °C":
    def cool_full_dry_room(x,y):
        p00 =       40.67
        p10 =     0.01687
        p01 =      0.6638
        p20 =    0.006678
        p11 =    -0.02612
        p02 =      0.1324
        p30 =   0.0001946
        p21 =  -0.0004645
        p12 =    0.009223
        p03 =    -0.03712
        p40 =  -5.606e-06
        p31 =   2.382e-05
        p22 =  -0.0001805
        p13 =   3.793e-06
        p04 =    0.002084
        return(p00 + p10*x + p01*y + p20*x**2 + p11*x*y + p02*y**2 + p30*x**3 + p21*x**2*y + p12*x*y**2 + p03*y**3 + p40*x**4 + p31*x**3*y + p22*x**2*y**2 + p13*x*y**3 + p04*y**4)

if dew_point == "-40 °C":
    def cool_full_dry_room(x, y): 
        return ( 42.23 + 0.00639*x + 0.2352*y + 0.002451*x**2 - 0.009739*x*y + 0.04956*y**2 + 7.133e-05*x**3 - 0.000167*x**2*y + 0.003407*x*y**2 - 0.01413*y**3 - 2.049e-06*x**4 + 8.536e-06*x**3*y - 6.456e-05*x**2*y**2 - 7.002e-06*x*y**3 + 0.0007891*y**4 )
# 2Rotor-System

#-----FULL ELECTRIC DRY ROOM--------------------------------------
if dew_point == "-60 °C":
    def electr_full_dry_room():
        return(72.5)

if dew_point == "-50 °C":
    def electr_full_dry_room():
        return (22.58)

if dew_point == "-40 °C":
    def electr_full_dry_room():
        return(5.26)


#-----TEILLAST------------------------------------------------------
def cool_partial_dry_room(x,y):
    return(x)



#------------------------ENERGIEKONZEPTE--------------------------------------------------------
#-----WIRKUNGSGRADE--------------------------------------------------------------
#-----Carnot-Formel 85 Grad Vorlauf----------------------------------
def cop(K):
    t_h=358.15
    real_faktor=0.625
    if K>(t_h+0.2):
        return 1e3
    else:
        n_c=t_h/(t_h-K)
        n_c_real=n_c*real_faktor
        return(n_c_real)
    
#-----EERT Tischrückkühler trocken-----------------------------------
def eert(K):
    if K>312.15:
        return 14
    else:
        e=-24.067*K + 7526.4
        return(e)



#-----BERECHNUNG ENDLAST-------------------------------------------------------
#-----Kälteerzeugung KKM
def strom_eert(e, kaelte):
    cop_kkm = 6.1
    end_kaelteleistung = kaelte/cop_kkm + (kaelte*1.3)/e
    return(end_kaelteleistung)

#-----Konzept 1 - Brennwertkessel---------------------------------------
def brennwertkessel_wirkungsgrad(heat):
    n = 0.965
    end_waermelast = heat/n
    return(end_waermelast)

#-----Konzept 2 - BHKW--------------------------------------------------
def bhkw_w_wirkungsgrad(x):
    n = 0.55
    return x/n

def bhkw_s_wirkungsgrad(x):
    n = 0.35
    return x*n

def bhkw_ges_wirkungsgrad(x):
    n=0.9
    return x/n
#-----Konzept 3 - Heat Pump--------------------------------------------
def strom_cop(n_c, waerme):
    end_waermeleistung = waerme/n_c
    return(end_waermeleistung)

#-----Konzept 4 - Kombi-WP----------------------------------------------
def kombi_wp_w_end(x):
    cop_kwp=5.7396
    return x/cop_kwp

def kombi_wp_k_end(x):
    cop_kkm=6.1
    return x/cop_kkm


#---------------------DURCHLAUFEN DER FORMELN ZUR BESTIMMUNG DES LASTVERLAUFS------------------------------------------------------

for i in range(len(t)):
    
    # Extract a pair of values for each iteration
    temp_val = t[i]
    rhum_val = rhum[i]
    pres_val = pres[i]
    t_k = t_kelvin[i]
    
    #Absolute Feuchte berechnen
    hum_abs_val = hum_abs(temp_val, rhum_val, pres_val)
    f_abs.append(hum_abs_val)
    
    #EERT-Wert berechnen
    eert_val = eert(t_k)
    eert_data.append(eert_val)
    
    #COP-Wert berechnen
    cop_val = cop(t_k)
    cop_data.append(cop_val)
    
    #Kälteleistung berechnen
    cool_val = cool_full_dry_room(temp_val, hum_abs_val)
    cool_data.append(cool_val)
    
    #Wärmeleistung berechnen
    heat_val = heat_full_dry_room(temp_val, hum_abs_val)
    heat_data.append(heat_val)
    
    #partial_heat_val = dryroom_heat_partial(temp_val, hum_abs_val)
    #heat_data_partial.append(partial_heat_val)
    
    #cool_val_partial = dryroom_cool_partial(temp_val, hum_abs_val)
    #heat_data_partial.append(cool_val_partial)
    
    #End-Kühlleistung berechnen
    strom_wp_k_end.append(strom_eert(eert_val, cool_val))
    
    #Heat Pump
    #End-Wärmeleistung berechnen
    strom_wp_w_end.append(strom_cop(cop_val, heat_val))
    
    #Brennwertkessel
    brennstoff_w_end.append(brennwertkessel_wirkungsgrad(heat_val))
    
    #Strom
    strom_electr_end.append(electr_full_dry_room())



#-----Durchschnittswerte der Effizienz--------------------------------------------------------------------------------
cop_avg = mean(cop_data)
eer_avg = mean(eert_data)
cop_kkm = 6.1

#---------------------BERECHNUNG DER ENDLAST ZUM GESAMTENERGIEBEDARF----------------------------------------------------------------
#-----REIN_ UND TROCKENRAUM--------------------------------------------------------------------------------------------
MA_factor = (MA_nach_Automatisierungsgrad(MA_in_RuT(production_capacity, cell_format))/2)
#-----Kälte RuT gesamt ausgeben-------------------------------------------------------------------------
RuT_GWh_k_nutz = sum(cool_data)/10**6 * MA_factor*production_day_factor

if energy_concept == 'Hybrid Heat Pump':
    RuT_GWh_k_end = kombi_wp_k_end(RuT_GWh_k_nutz)
else:
    RuT_GWh_k_end = sum(strom_wp_k_end)/10**6 * MA_factor*production_day_factor



#-----Wärme RuT gesamt ausgeben--------------------------------------------------------------------------
RuT_GWh_w_nutz = sum(heat_data)/10**6 * MA_factor *production_day_factor

if energy_concept == 'Natural Gas Boiler':
    RuT_GWh_w_end = sum(brennstoff_w_end)/10**6 * MA_factor *production_day_factor
if energy_concept == 'Cogeneration Unit':
    RuT_GWh_w_end = bhkw_w_wirkungsgrad(RuT_GWh_w_nutz)
if energy_concept == 'Heat Pump':
    RuT_GWh_w_end = sum(strom_wp_w_end)/10**6 * MA_factor *production_day_factor
if energy_concept == 'Hybrid Heat Pump':
    RuT_GWh_w_end = kombi_wp_w_end(RuT_GWh_w_nutz)


#-----Strom RuT gesamt ausgeben--------------------------------------------------------------------------
RuT_GWh_s_nutz = sum(strom_electr_end)/10**6 * (MA_nach_Automatisierungsgrad(MA_in_RuT(production_capacity, cell_format))/2)*production_day_factor

if energy_concept == 'Cogeneration Unit':
    RuT_GWh_s_end = RuT_GWh_s_nutz - bhkw_s_wirkungsgrad(RuT_GWh_s_nutz)
else:
    RuT_GWh_s_end = RuT_GWh_s_nutz

#------Gesamtenergiemenge RuT ausgeben-------------------------------------------------------------------
RuT_GWh_kum_ges = (RuT_GWh_k_end + RuT_GWh_w_end + RuT_GWh_s_end)






#------GEBÄUDETECHNIK------------------------------------------------------------------------------------------------
#-----Nutzlast-------------------------------------------------------------------------------------------
#-----       (INFO: RLT unabhängig von production days)       -------------------
RLT_GWh_k_nutz = RLT_Kaeltelast(production_capacity)
RLT_GWh_w_nutz = RLT_Waermelast(production_capacity)
RLT_GWh_s_nutz = RLT_Stromlast(production_capacity)

#-----Endwärme---------------------------------------------------------------------
if energy_concept == 'Hybrid Heat Pump':
    RLT_GWh_k_end = kombi_wp_k_end(RLT_GWh_k_nutz)
else:
    RLT_GWh_k_end = strom_eert(eer_avg, RLT_GWh_k_nutz)

#-----Endkälte---------------------------------------------------------------------
if energy_concept == 'Natural Gas Boiler':
    RLT_GWh_w_end = brennwertkessel_wirkungsgrad(RLT_GWh_w_nutz)
if energy_concept == 'Cogeneration Unit':
    RLT_GWh_w_end = bhkw_w_wirkungsgrad(RLT_GWh_w_nutz)
if energy_concept == 'Heat Pump':
    RLT_GWh_w_end = RLT_GWh_w_nutz/cop_avg
if energy_concept == 'Hybrid Heat Pump':
    RLT_GWh_w_end = kombi_wp_w_end(RLT_GWh_w_nutz)

#-----Endstrom-------------------------------------------------------------------
if energy_concept == 'Cogeneration Unit':
    RLT_GWh_s_end = RLT_GWh_s_nutz - bhkw_s_wirkungsgrad(RLT_GWh_s_nutz)
else:
    RLT_GWh_s_end = RLT_GWh_s_nutz




#-----PROZESSE-------------------------------------------------------------------------------------------
#-----Nutzlast---------------------------------------------------------------------
PRO_GWh_k_nutz = Prozess_Kaeltenutzlast(production_capacity)*production_day_factor
PRO_GWh_s_nutz = Prozess_Stromnutzlast(production_capacity)*production_day_factor

#-----Endkälte--------------------------------------------------------------------
if energy_concept == 'Hybrid Heat Pump':
    PRO_GWh_k_end = kombi_wp_k_end(PRO_GWh_k_nutz)
else:
    PRO_GWh_k_end = strom_eert(eer_avg, PRO_GWh_k_nutz)


#-----Endstrom-------------------------------------------------------------------
if energy_concept == 'Cogeneration Unit':
    PRO_GWh_s_end = PRO_GWh_s_nutz - bhkw_s_wirkungsgrad(RLT_GWh_s_nutz)
else:
    PRO_GWh_s_end = PRO_GWh_s_nutz





#-----------GESAMTFACTORY ENERGIEVERBRÄUCHE---------------------------------------------------------------------------
#-----Nutzlast Gesamtfabrik---------------------------------------------------
gesamtfabrik_k_nutz = (RuT_GWh_k_nutz+PRO_GWh_k_nutz+RLT_GWh_k_nutz)
gesamtfabrik_w_nutz = (RuT_GWh_w_nutz+RLT_GWh_w_nutz)
gesamtfabrik_s_nutz = (RuT_GWh_s_nutz+PRO_GWh_s_nutz+RLT_GWh_s_nutz)

gesamtfabrik_ges_nutz = gesamtfabrik_k_nutz + gesamtfabrik_w_nutz + gesamtfabrik_s_nutz

#-----Endlast Gesamtfabrik----------------------------------------------------
gesamtfabrik_k_end = (RuT_GWh_k_end+PRO_GWh_k_end+RLT_GWh_k_end)
gesamtfabrik_w_end = (RuT_GWh_w_end+RLT_GWh_w_end)
gesamtfabrik_s_end = (RuT_GWh_s_end+PRO_GWh_s_end+RLT_GWh_s_end)

gesamtfabrik_ges_end = gesamtfabrik_k_end + gesamtfabrik_w_end + gesamtfabrik_s_end

#-----Gesamtfabrik Daten & Werte----------------------------------------------
energiefaktor = gesamtfabrik_ges_end/(production_capacity*production_day_factor_315)


#-----Emissionen--------------------------------------------------------------------------------
def co2_natual_gas(x):
    kg_co2_GWh = 0.24 * 10**6
    return(kg_co2_GWh * x)

#-----Strom-Emissionen des Strommix nach country_code-------------------------
# Fetch the data.
df_co2 = pd.read_csv("co2_emission_factors.csv") #emissions in kg/kWh

# CO2-Emissionen abrufen
co2_emissions = df_co2[df_co2["Code"] == country_code.upper()]["emissions"] #upper() um country_code in Großbuchstaben umzuwandeln

# In Streamlit anzeigen


def co2_electric(x):
    kg_co2_GWh = 10**6 * co2_emissions
    return(x*kg_co2_GWh)


if energy_concept=="Cogeneration Unit":
    natural_gas_usage=bhkw_ges_wirkungsgrad(gesamtfabrik_w_end+bhkw_s_wirkungsgrad(gesamtfabrik_s_nutz))
elif energy_concept=="Natural Gas Boiler":
    natural_gas_usage=gesamtfabrik_w_end
else:
    natural_gas_usage=0

#-----natural gas in m³------------------------------------------
mio_cubic_meters = ((10**6 * natural_gas_usage)/11) / 10**6
mio_cubic_meters_daily = mio_cubic_meters/365


if energy_concept=="Cogeneration Unit":
    electricity_usage=gesamtfabrik_ges_end-natural_gas_usage
elif energy_concept=="Natural Gas Boiler":
    electricity_usage=gesamtfabrik_ges_end-gesamtfabrik_w_end
else:
    electricity_usage=gesamtfabrik_ges_end



#-----TOTAl------
natural_gas_emissions_kilotons=(co2_electric(electricity_usage)+co2_natual_gas(natural_gas_usage))/10**6 


#--------------------DATA VISUALIZER-----------------------------------------------------------------------------------
#-----Differenzen------------------------------------------
if cell_format == "Pouch":
    avg_energiefaktor=45
if cell_format == "Cylindrical":
    avg_energiefaktor=50
if cell_format == "Prismatic":
    avg_energiefaktor=55
dif_energiefaktor=(1-(energiefaktor/avg_energiefaktor))*100


#-----Row A-----------------------------------------------------------------
container_a = st.container(border=True)
with container_a:
    st.subheader(":material/key: Key Values")
    a1, a2, a3, a4= st.columns(4)
    a1.metric(":material/energy_program_time_used: Energy Factor [kWh/kWhcell]", f"{round(energiefaktor,2)}", delta=f"{round(dif_energiefaktor, 1)} %" )
    a2.metric(":material/bolt: Estimated Connection power",f"{round(((((gesamtfabrik_ges_end-natural_gas_usage)/8760)*1.2)*10**3),2)} MW")
    a3.metric(":material/power: Electricity input [GWh/a]",round((gesamtfabrik_ges_end-natural_gas_usage),2))
    a4.metric(":material/water_drop: Natural Gas input [mio. m³/a]", round(mio_cubic_meters,2) )

#-----Row B-----------------------------------------------------------------
container_b = st.container(border=True)
with container_b:
    st.subheader(":material/energy_program_time_used: Overall Energy Usage by type")
    b1, b2, b3, b4 = st.columns(4)
    b1.metric(":material/heat: Heat energy output [GWh/a]",round(gesamtfabrik_w_nutz,2))
    b2.metric(":material/mode_cool: Cooling energy output [GWh/a]",round(gesamtfabrik_k_nutz,2))
    b3.metric(":material/bolt: Electrical energy output [GWh/a]",round(gesamtfabrik_s_nutz,2))
    b4.metric("Total energy output [GWh/a]", f"{round(gesamtfabrik_ges_nutz,2)}")

#-----Row B-----------------------------------------------------------------
container_b = st.container(border=True)
with container_b:
    st.subheader(":material/energy_program_time_used: Useful Energy Factors by type ")
    b1, b2, b3, b4 = st.columns(4)
    b1.metric(":material/heat: Heat energy factor [kWh/kWhcell]",round((gesamtfabrik_w_nutz/(production_capacity*production_day_factor_315)),2))
    b2.metric(":material/mode_cool: Cooling energy factor [kWh/kWhcell]",round((gesamtfabrik_k_nutz/(production_capacity*production_day_factor_315)),2))
    b3.metric(":material/bolt: Electrical energy factor [kWh/kWhcell]",round((gesamtfabrik_s_nutz/(production_capacity*production_day_factor_315)),2))
    b4.metric(":material/energy_program_time_used: Total Useful Energy Factor [kWh/kWhcell]", f"{round((gesamtfabrik_ges_nutz/(production_capacity*production_day_factor_315)),2)}")

#------dry room extras----------------------------------------------------
container_b = st.container(border=True)
with container_b:
    st.subheader(":material/cool_to_dry: Dry Room Energy Usage")
    b1, b2, b3, b4 = st.columns(4)
    b1.metric(":material/heat: Heat energy usage [GWh/a]",round(RuT_GWh_w_nutz,2))
    b2.metric(":material/mode_cool: Cooling energy usage [GWh/a]",round(RuT_GWh_k_nutz,2))
    b3.metric(":material/bolt: Electrical energy usage [GWh/a]",round(RuT_GWh_s_nutz,2))
    b4.metric(":material/input: Total energy input [GWh/a]", f"{round((RuT_GWh_w_end+RuT_GWh_k_end+RuT_GWh_s_end),2)}")
    
#-----row b2---------------------------------------------------------------
container_c = st.container(border=True)
with container_c:
    st.subheader(":material/analytics: Additional Information")
    b5, b6, b7, b8 = st.columns(4)
    b5.metric(":material/eco: CO2-emissions [kilotons/year]",round((natural_gas_emissions_kilotons),1), help="This metric considers the CO2 emissions from both your local electricity supply and natural gas consumption.")
    b6.metric(":material/eco: CO2-emissions factor [kg/kWh]",round((natural_gas_emissions_kilotons/(production_capacity*production_day_factor_315)),2), help="This metric considers the CO2 emissions from both your local electricity supply and natural gas consumption.")
    b7.metric("Total Electricity Costs [Mio.€/GWh]", round(((((electricity_usage*10**6)*electricity_price)/(production_capacity*production_day_factor_315))/10**6),2))  
    b8.metric(":material/groups: People in Dry Rooms",int(MA_nach_Automatisierungsgrad((MA_in_RuT(production_capacity, cell_format)))))

#Excel-Export-------------
def to_excel_openpyxl(df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Daten in das Worksheet schreiben
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Formatierungen definieren
    number_format = '0.00'
    bold_format = Font(bold=True)
    header_format = Font(bold=True, size=14)
    left_border = Border(left=Side(style='thick'))
    left_border_thin = Border(left=Side(style='thin'))
    bottom_border = Border(bottom=Side(style='thick'))
    top_border = Border(top=Side(style='thick'))
    orange_fill = PatternFill(start_color='F26B43', end_color='F26B43', fill_type='solid')
    blue1_fill = PatternFill(start_color='03738C', end_color='03738C', fill_type='solid')
    blue2_fill = PatternFill(start_color='0396A6', end_color='0396A6', fill_type='solid')
    blue3_fill = PatternFill(start_color='08A696', end_color='08A696', fill_type='solid')

    # Formatierungen anwenden
    for row in ws.iter_rows():
        for cell in row:
            if cell.column_letter == 'A':
                cell.number_format = number_format
                cell.font = bold_format
            if cell.column_letter == 'B':
                cell.number_format = number_format
                cell.border = left_border_thin
            if cell.column_letter == 'C':
                cell.font = bold_format
                cell.border = left_border
            if cell.column_letter == 'D':
                cell.number_format = number_format
                cell.border = left_border_thin
            if cell.column_letter == 'E':
                cell.font = bold_format
                cell.border = left_border
            if cell.column_letter == 'F':
                cell.number_format = number_format
                cell.border = left_border_thin
            if cell.column_letter == 'G':
                cell.font = bold_format
                cell.border = left_border
            if cell.column_letter == 'H':
                cell.number_format = number_format
                cell.border = left_border_thin
            if cell.column_letter == 'I':
                cell.number_format = number_format
                cell.border = left_border

    # Spaltenbreiten festlegen
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['G'].width = 30

    # Zellen verbinden - Überschriften
    ws.merge_cells('A1:B1')
    ws['A1'] = 'YOUR FACTORY'
    ws['A1'].font = header_format
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = orange_fill

    ws.merge_cells('C1:D1')
    ws['C1'] = 'Key Values'
    ws['C1'].font = header_format
    ws['C1'].alignment = Alignment(horizontal='center')
    ws['C1'].fill = blue1_fill

    ws.merge_cells('E1:F1')
    ws['E1'] = 'Overall Energy Usage by type'
    ws['E1'].font = header_format
    ws['E1'].alignment = Alignment(horizontal='center')
    ws['E1'].fill = blue2_fill

    ws.merge_cells('G1:H1')
    ws['G1'] = 'Additional Information'
    ws['G1'].font = header_format
    ws['G1'].alignment = Alignment(horizontal='center')
    ws['G1'].fill = blue3_fill

    # ... borders ...
    for row in ws['A1:H1']:
        for cell in row:
            cell.border = bottom_border

    for row in ws['A7:H7']:
        for cell in row:
            cell.border = top_border
    
    wb.save(output)
    processed_data = output.getvalue()
    return processed_data
# Beispiel DataFrame (ersetze dies mit deinen eigenen Daten)

data_excel = {
  "YOUR FACTORY": ["Location", "Cell Format", "Production Capacity [GWh/a]", "Production Days", "Dew Point [°C]"],
  "-": [location.address, cell_format, production_capacity, production_days, dew_point],
  "Key Values": ["Energy Factor", "Estimated Connection Power ", "Electricity Input [GWh/a]", "Natural Gas Input [mio. m³/a]", ""],
  "--": [energiefaktor, ((((gesamtfabrik_ges_end-natural_gas_usage)/8760)*1.2)*10**3), gesamtfabrik_ges_end-natural_gas_usage, mio_cubic_meters, ""],
  "Overall Energy Usage by type": ["Heat energy output [GWh/a]", "Cooling energy output [GWh/a]", "Electrical energy output [GWh/a]", "Total energy output [GWh/a]", ""],
  "---": [gesamtfabrik_w_nutz, gesamtfabrik_k_nutz, gesamtfabrik_s_nutz, gesamtfabrik_ges_nutz, ""],
  "Additional Information": [" CO2-emissions [kilotons/year]", " CO2-emissions factor [kg/kWh]", "Total Electricity Costs [mio.€/GWh]", " People in Dry Rooms", ""],
  "----": [float(natural_gas_emissions_kilotons), float(natural_gas_emissions_kilotons/(production_capacity*production_day_factor_315)), ((((electricity_usage*10**6)*electricity_price)/(production_capacity*production_day_factor_315))/10**6), int(MA_nach_Automatisierungsgrad((MA_in_RuT(production_capacity, cell_format)))), ""],
}

df_excel = pd.DataFrame(data_excel)  # DataFrame erstellen
df_xlsx = to_excel_openpyxl(df_excel)
st.title("Export to Excel")
st.download_button(label='📥 Download Current Results',
                                data=df_xlsx ,
                                file_name= f"Gigafactory_Builder_Export_{city}_{production_capacity}GWh_{cell_format}.xlsx")


#---------Row D - SANKEY DIAGRAM-----------------------------------------------------
#-----define sankey states-------------------------------------------------------
if energy_concept == "Natural Gas Boiler":
        df = pd.DataFrame(
            {
                "source": ["Electricity", "Electricity", "Natural Gas", "Grid Connection", "Natural Gas Boiler", "Cooling Unit", "Grid Connection", "Natural Gas Boiler", "Cooling Unit", "Grid Connection", "Natural Gas Boiler", "Cooling Unit"],
                "target": ["Grid Connection", "Cooling Unit", "Natural Gas Boiler", "Manufacturing", "Manufacturing", "Manufacturing", "Dry Room", "Dry Room","Dry Room", "Building", "Building", "Building"],
                "value": [gesamtfabrik_s_end, gesamtfabrik_k_end, gesamtfabrik_w_end, PRO_GWh_s_nutz, 0, PRO_GWh_k_nutz, RuT_GWh_s_nutz, RuT_GWh_w_nutz, RuT_GWh_k_nutz, RLT_GWh_s_nutz, RLT_GWh_w_nutz, RLT_GWh_k_nutz],
            }
        )
if energy_concept == "Cogeneration Unit":
        df = pd.DataFrame(
            {
                "source": ["Electricity", "Electricity", "Natural Gas","Cogeneration Unit", "Cogeneration Unit", "Grid Connection", "Grid Connection","Grid Connection", "Cooling Unit", "Cooling Unit", "Cooling Unit", "Boiler", "Boiler" ],
                "target": ["Grid Connection", "Cooling Unit", "Cogeneration Unit", "Grid Connection", "Boiler", "Manufacturing", "Dry Room", "Building", "Manufacturing", "Dry Room", "Building", "Dry Room", "Building"],
                "value": [gesamtfabrik_s_end, gesamtfabrik_k_end, natural_gas_usage, (bhkw_s_wirkungsgrad(gesamtfabrik_s_nutz)), gesamtfabrik_w_end, PRO_GWh_s_nutz, RuT_GWh_s_nutz, RLT_GWh_s_nutz, PRO_GWh_k_nutz, RuT_GWh_k_nutz, RLT_GWh_k_nutz, RuT_GWh_w_nutz, RLT_GWh_w_nutz],
            }
        )
if energy_concept == "Heat Pump":
        df = pd.DataFrame(
            {
                "source": ["Electricity", "Manufacturing", "Manufacturing","Electricity", "Building", "Building", "Building","Electricity", "Dry Room","Dry Room", "Dry Room"],
                "target": ["Manufacturing", "Grid Connection", "Cooling Unit", "Building", "Cooling Unit", "Grid Connection", "Heat Pump", "Dry Room", "Cooling Unit", "Grid Connection", "Heat Pump"],
                "value": [(PRO_GWh_s_end+PRO_GWh_k_end), PRO_GWh_s_nutz, PRO_GWh_k_nutz, (RLT_GWh_k_end+RLT_GWh_s_end+RLT_GWh_w_end), RLT_GWh_k_nutz, RLT_GWh_s_nutz, RLT_GWh_w_nutz, (RuT_GWh_k_end+RuT_GWh_s_end+RuT_GWh_w_end), RuT_GWh_k_nutz, RuT_GWh_s_nutz, RuT_GWh_w_nutz],
            }
        )
        
if energy_concept == "Hybrid Heat Pump":
        df = pd.DataFrame(
            {
                "source": ["Electricity", "Manufacturing", "Manufacturing","Electricity", "Building", "Building", "Building","Electricity", "Dry Room","Dry Room", "Dry Room"],
                "target": ["Manufacturing", "Grid Connection", "Cooling Unit", "Building", "Cooling Unit", "Grid Connection", "Hybrid Heat Pump", "Dry Room", "Cooling Unit", "Grid Connection", "Hybrid Heat Pump"],
                "value": [(PRO_GWh_s_end+PRO_GWh_k_end), PRO_GWh_s_nutz, PRO_GWh_k_nutz, (RLT_GWh_k_end+RLT_GWh_s_end+RLT_GWh_w_end), RLT_GWh_k_nutz, RLT_GWh_s_nutz, RLT_GWh_w_nutz, (RuT_GWh_k_end+RuT_GWh_s_end+RuT_GWh_w_end), RuT_GWh_k_nutz, RuT_GWh_s_nutz, RuT_GWh_w_nutz],
            }
        )


def draw_sankey(df):
    flows = list(df[["source", "target", "value"]].itertuples(index=False, name=None))
    # remove empty and nan values
    flows_clean = [x for x in flows if x[0] and x[1] and x[2] > 0]
    
    diagram = Sankey(
        flows=flows_clean,
        cmap=plt.get_cmap("Pastel1"),
        flow_color_mode="source",
        node_opts={"label_opts": {"fontsize": 7, "fontname": "monospace"}},  # Set font to monospace
        flow_opts={"curvature": 8/10},
    )
    _, col2, _ = st.columns([0.1, 5, 0.5])
    with col2:
        diagram.draw()
        st.pyplot(plt)
        img = io.BytesIO()
        plt.savefig(img, format="png")
        st.session_state.image = img
 
 
def empty_df():
    df = pd.DataFrame({"source": [""], "target": [""], "value": [None]})
    st.session_state.df = df.astype({"value": float})
 
 
 
sankey_placeholder = st.empty()

#-----draw sankey plot ---------------------------------------------------
container_sankey = st.container(border=True)
with container_sankey:
    sankey1, sankey2 = st.columns([7,3])
    with sankey1:
        st.header(":material/account_tree: Sankey-Plot", help="This Sankey plot is still work in progress, only the plot shown when choosing the Natural Gas Boiler as your energy concept is done. The others still have some improvements coming. :D")
    with sankey2:
        st.subheader("all values in GWh/a")
    draw_sankey(df)
    

end1, end2 = st.columns([7,3])
#with end1:
    #with st.container(border=True):
        #st.image("Gigafactory Builder Logo.png")
with end2:
    st.markdown("`Created by Fraunhofer FFB`")
#--------------------------by tarek lichtenfeld, december 2024------------------------------------